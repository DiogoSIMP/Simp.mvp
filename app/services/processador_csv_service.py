import pandas as pd
import glob
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.models.database import get_db_connection, formatar_nome

class ProcessadorCSVService:
    
    def __init__(self):
        self.PERCENTUAL_PAGAMENTO = 0.6
        self.DESCONTO_FIXO = 0.35
    
    def classificar_tipo(self, desc):
        """Classifica o tipo de valor pela descrição (seguindo a lógica do arquivo de referência)."""
        desc = str(desc).lower()
        if "gorjeta" in desc:
            return "gorjeta"
        elif "promocao entregador" in desc:
            return "promo"
        elif "corridas concluidas" in desc:
            return "corridas"
        elif "valor por hora online" in desc:
            return "online_time"
        elif "route_with_occurrence" in desc:
            return "rotas_com_ocorrencia"
        elif "tempo de espera na origem" in desc:
            # tempo_espera é classificado mas NÃO entra no valor_total (seguindo arquivo de referência)
            return "tempo_espera"
        else:
            return "outros"
    
    def processar_csv(self, caminho_arquivo):
        """Processa um arquivo CSV individual"""
        try:
            print(f"📂 Processando arquivo: {os.path.basename(caminho_arquivo)}")
            
            # Tentar diferentes encodings comuns para arquivos brasileiros
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
            df = None
            encoding_usado = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(caminho_arquivo, sep=";", encoding=encoding)
                    encoding_usado = encoding
                    print(f"   ✅ Arquivo lido com encoding: {encoding}")
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            if df is None:
                raise Exception("Não foi possível determinar a codificação do arquivo. Tente converter para UTF-8.")
            
            print(f"   📊 Colunas encontradas: {list(df.columns)}")
            print(f"   📈 Linhas: {len(df)}")
            
            # Validar colunas necessárias
            colunas_necessarias = [
                "id_da_pessoa_entregadora", "recebedor", "subpraca", "praca", 
                "valor", "descricao", "periodo"
            ]
            
            for coluna in colunas_necessarias:
                if coluna not in df.columns:
                    raise ValueError(f"Coluna obrigatória ausente: {coluna}")
            
            # Processar dados - PRESERVAR descrição original antes de converter para lowercase
            df["descricao_original"] = df["descricao"].astype(str).fillna("")
            df["descricao"] = df["descricao_original"].str.lower().fillna("")
            
            # Converter valores: verificar se já é numérico ou se precisa converter formato brasileiro
            if df["valor"].dtype in ['float64', 'int64', 'float', 'int']:
                # Já é numérico (formato internacional com ponto decimal)
                df["valor"] = pd.to_numeric(df["valor"], errors='coerce').fillna(0.0)
            else:
                # É string - pode estar no formato brasileiro (vírgula como separador decimal)
                # Verificar se há vírgula (formato brasileiro) ou se é só número
                df["valor"] = (
                    df["valor"].astype(str)
                    .str.replace(".", "", regex=False)  # Remove separador de milhar
                    .str.replace(",", ".", regex=False)  # Converte vírgula para ponto decimal
                    .astype(float)
                )
            
            # Classificar tipo de valor - TODAS as ocorrências serão classificadas
            df["tipo_valor"] = df["descricao"].apply(self.classificar_tipo)
            
            # Verificar se há valores NaN ou vazios que não foram classificados
            df["tipo_valor"] = df["tipo_valor"].fillna("outros")
            
            # Garantir que valores NaN sejam 0
            df["valor"] = df["valor"].fillna(0.0)
            
            print(f"   ✅ Arquivo processado com sucesso")
            return df
            
        except Exception as e:
            print(f"   ❌ Erro no processamento: {str(e)}")
            raise Exception(f"Erro ao processar arquivo {os.path.basename(caminho_arquivo)}: {str(e)}")
    
    def consolidar_entregadores(self, df, data_filtro=None):
        """
        Consolida dados de pagamento por entregador
        
        Args:
            df: DataFrame com os dados
            data_filtro: Data no formato date para filtrar apenas dados dessa data (opcional)
        """
        try:
            # Verificar se o DataFrame não está vazio
            if df.empty:
                return pd.DataFrame()
            
            # Filtrar por data do período de referência se especificado
            if data_filtro:
                colunas_data = ['data_do_periodo_de_referencia', 'data_periodo', 'data_referencia', 'periodo_data']
                coluna_data_encontrada = None
                
                for col in colunas_data:
                    if col in df.columns:
                        coluna_data_encontrada = col
                        break
                
                if coluna_data_encontrada:
                    try:
                        # Converter coluna para data e filtrar
                        df[coluna_data_encontrada] = pd.to_datetime(df[coluna_data_encontrada], errors='coerce').dt.date
                        df_antes = len(df)
                        df = df[df[coluna_data_encontrada] == data_filtro]
                        df_depois = len(df)
                        print(f"   📅 Filtrado por data {data_filtro}: {df_antes} → {df_depois} linhas")
                    except Exception as e:
                        print(f"   ⚠️  Erro ao filtrar por data na consolidação: {str(e)}")
            
            # Verificar se há descrições não reconhecidas (classificadas como "outros")
            # Agrupar descrições originais para "outros" para referência
            # Usar "descricao_original" se existir, senão usar "descricao"
            col_descricao = "descricao_original" if "descricao_original" in df.columns else "descricao"
            outros_descricoes = df[df["tipo_valor"] == "outros"].groupby("id_da_pessoa_entregadora")[col_descricao].apply(
                lambda x: " | ".join(sorted(set(x.dropna().astype(str)))) if len(x.dropna()) > 0 else ""
            ).reset_index()
            outros_descricoes.columns = ["id_da_pessoa_entregadora", "descricoes_outros"]
            
            # Pivot table para tipos de valores - SOMAR TODAS as ocorrências
            pivot = (
                df.pivot_table(
                    index=["id_da_pessoa_entregadora", "recebedor"],
                    values="valor",
                    columns="tipo_valor",
                    aggfunc="sum",  # Soma todas as ocorrências
                    fill_value=0
                )
                .reset_index()
            )
            
            # Adicionar descrições de "outros" ao pivot
            if not outros_descricoes.empty:
                pivot = pd.merge(pivot, outros_descricoes, on="id_da_pessoa_entregadora", how="left")
                pivot["descricoes_outros"] = pivot["descricoes_outros"].fillna("")
            else:
                pivot["descricoes_outros"] = ""
            
            # Informações de praças
            sub_data = (
                df.groupby("id_da_pessoa_entregadora", as_index=False)
                .agg(
                    subpracas=("subpraca", lambda x: " / ".join(sorted(set(x.dropna().astype(str)))) if len(x.dropna()) > 0 else ""),
                    qtd_subpracas=("subpraca", lambda x: len(set(x.dropna()))),
                    pracas=("praca", lambda x: " / ".join(sorted(set(x.dropna().astype(str)))) if len(x.dropna()) > 0 else ""),
                    qtd_pracas=("praca", lambda x: len(set(x.dropna())))
                )
            )
            
            consolidado = pd.merge(pivot, sub_data, on="id_da_pessoa_entregadora", how="left")
            
            # Garantir que subpracas e pracas não sejam NaN
            consolidado['subpracas'] = consolidado['subpracas'].fillna("").astype(str)
            consolidado['pracas'] = consolidado['pracas'].fillna("").astype(str)
            
            # Tipos de valor que entram no cálculo
            # tempo_espera TAMBÉM entra no valor_total
            tipos_valor_calculo = ["corridas", "gorjeta", "promo", "online_time", "rotas_com_ocorrencia", "tempo_espera", "outros"]
            tipos_valor_exibicao = tipos_valor_calculo  # Todos os tipos são exibidos
            
            # Garantir que todas as colunas existam
            for tipo in tipos_valor_exibicao:
                if tipo not in consolidado.columns:
                    consolidado[tipo] = 0.0
            
            # Garantir que todos os valores sejam numéricos
            for col in tipos_valor_exibicao:
                consolidado[col] = pd.to_numeric(consolidado[col], errors='coerce').fillna(0.0)
            
            # Calcular valor total
            # Soma todos os tipos que entram no cálculo (incluindo tempo_espera)
            consolidado["valor_total"] = (
                consolidado.get("corridas", 0)
                + consolidado.get("gorjeta", 0)
                + consolidado.get("promo", 0)
                + consolidado.get("online_time", 0)
                + consolidado.get("rotas_com_ocorrencia", 0)
                + consolidado.get("tempo_espera", 0)
                + consolidado.get("outros", 0)
            )
            
            # Arredondar valores individuais (seguindo arquivo de referência)
            for c in tipos_valor_calculo + ["valor_total"]:
                consolidado[c] = consolidado[c].round(2)
            
            # Cálculo do 60% (seguindo exatamente a lógica do arquivo de referência)
            # Base: valor_total - gorjeta (gorjeta não entra no cálculo do adiantamento)
            base_calculo = consolidado["valor_total"] - consolidado.get("gorjeta", 0)
            consolidado["valor_60_percent"] = (base_calculo * self.PERCENTUAL_PAGAMENTO).round(2)
            
            # Valor final: 60% menos desconto fixo de R$ 0.35
            consolidado["valor_final"] = (consolidado["valor_60_percent"] - self.DESCONTO_FIXO).round(2)
            
            # Garantir que valores não sejam negativos
            consolidado["valor_60_percent"] = consolidado["valor_60_percent"].clip(lower=0.0)
            consolidado["valor_final"] = consolidado["valor_final"].clip(lower=0.0)
            
            # Debug: verificar cálculos
            print("🔍 DEBUG - Cálculos:")
            print(f"   Valor Total: {consolidado['valor_total'].iloc[0]}")
            print(f"   Gorjetas: {consolidado['gorjeta'].iloc[0]}")
            print(f"   Base cálculo: {base_calculo.iloc[0]}")
            print(f"   60% da base: {consolidado['valor_60_percent'].iloc[0]}")
            print(f"   Valor final: {consolidado['valor_final'].iloc[0]}")
            
            return consolidado
            
        except Exception as e:
            print(f"Erro na consolidação: {str(e)}")
            return pd.DataFrame()
    
    def obter_detalhes_entregador(self, id_entregador):
        """Obtém detalhes de um entregador específico do banco de dados"""
        conn = get_db_connection()
        try:
            entregador = conn.execute('''
                SELECT * FROM entregadores WHERE id_da_pessoa_entregadora = ?
            ''', (id_entregador,)).fetchone()
            
            if entregador:
                return dict(entregador)
            return None
            
        except Exception as e:
            print(f"Erro ao buscar entregador: {str(e)}")
            return None
        finally:
            conn.close()
    
    def _obter_entregadores_cadastrados(self):
        """Obtém a lista de IDs dos entregadores cadastrados no banco"""
        from app.models.database import get_db_connection, is_postgresql_connection, get_db_cursor, get_db_placeholder
        
        conn = get_db_connection()
        cursor = get_db_cursor(conn)
        is_postgresql = is_postgresql_connection(conn)
        placeholder = get_db_placeholder(conn)
        
        try:
            if is_postgresql:
                cursor.execute(f"""
                    SELECT id_da_pessoa_entregadora 
                    FROM entregadores 
                    WHERE status = {placeholder}
                """, ('Ativo',))
            else:
                cursor.execute("""
                    SELECT id_da_pessoa_entregadora 
                    FROM entregadores 
                    WHERE status = ?
                """, ('Ativo',))
            
            resultados = cursor.fetchall()
            
            # Converter para lista de IDs
            ids = []
            for row in resultados:
                if is_postgresql:
                    ids.append(row['id_da_pessoa_entregadora'])
                else:
                    ids.append(row['id_da_pessoa_entregadora'])
            
            return ids
            
        except Exception as e:
            print(f"Erro ao buscar entregadores cadastrados: {str(e)}")
            return []
        finally:
            conn.close()
    
    
    def processar_multiplos_csv(self, lista_arquivos, data_filtro=None, ids_entregadores=None, filtrar_por_cadastrados=True):
        """
        Processa múltiplos arquivos CSV e retorna dados consolidados
        
        Args:
            lista_arquivos: Lista de caminhos dos arquivos CSV
            data_filtro: Data no formato YYYY-MM-DD para filtrar apenas CSVs desse dia (opcional)
            ids_entregadores: Lista de IDs de entregadores para filtrar (opcional)
            filtrar_por_cadastrados: Se True, filtra apenas entregadores cadastrados no banco. Se False, processa todos do CSV.
        """
        dataframes = []
        erros = []
        
        print(f"🔍 Processando {len(lista_arquivos)} arquivos...")  # DEBUG
        
        for arquivo in lista_arquivos:
            try:
                print(f"📂 Tentando processar: {os.path.basename(arquivo)}")  # DEBUG
                df = self.processar_csv(arquivo)
                
                # Se há filtro de data, filtrar pelos dados dentro do CSV
                # IMPORTANTE: Não filtrar pelo nome do arquivo, pois a data pode estar dentro dos dados
                if data_filtro and not df.empty:
                    # Verificar se há coluna de data do período de referência
                    colunas_data = ['data_do_periodo_de_referencia', 'data_periodo', 'data_referencia', 'periodo_data']
                    coluna_data_encontrada = None
                    
                    for col in colunas_data:
                        if col in df.columns:
                            coluna_data_encontrada = col
                            break
                    
                    if coluna_data_encontrada:
                        try:
                            # Converter coluna para data e filtrar
                            df[coluna_data_encontrada] = pd.to_datetime(df[coluna_data_encontrada], errors='coerce').dt.date
                            df = df[df[coluna_data_encontrada] == data_filtro]
                            print(f"   📅 Filtrado por data: {len(df)} linhas após filtro")
                        except Exception as e:
                            print(f"   ⚠️  Erro ao filtrar por data: {str(e)}")
                
                if not df.empty:
                    dataframes.append(df)
                    print(f"✅ Sucesso: {os.path.basename(arquivo)} - {len(df)} linhas")  # DEBUG
                else:
                    print(f"⏭️  Arquivo {os.path.basename(arquivo)} não tem dados para a data {data_filtro}")
            except Exception as e:
                erro_msg = f"Erro no arquivo {os.path.basename(arquivo)}: {str(e)}"
                erros.append(erro_msg)
                print(f"❌ {erro_msg}")  # DEBUG
        
        print(f"📊 Dataframes processados: {len(dataframes)}")  # DEBUG
        print(f"❌ Erros: {len(erros)}")  # DEBUG
        
        if len(dataframes) == 0:
            # Se há filtro de data, pode ser que nenhum arquivo tenha dados para aquela data
            # Retornar DataFrame vazio em vez de erro
            if data_filtro:
                print(f"⚠️  Nenhum arquivo contém dados para a data {data_filtro}")
                # Retornar DataFrame vazio com estrutura básica
                df_vazio = pd.DataFrame(columns=['id_da_pessoa_entregadora', 'recebedor', 'valor_total', 'valor_60_percent', 'valor_final'])
                return {
                    'df_completo': pd.DataFrame(),
                    'consolidado_geral': df_vazio,
                    'total_entregadores': 0,
                    'valor_total_geral': 0.0,
                    'data_processamento': datetime.now().strftime('%d/%m/%Y %H:%M'),
                    'erros': erros,
                    'total_arquivos': len(lista_arquivos),
                    'arquivos_sucesso': 0,
                    'arquivos_com_erro': len(erros),
                    'total_entregadores_cadastrados': 0,
                    'entregadores_com_dados': 0
                }
            else:
                raise Exception("Nenhum arquivo pôde ser processado")
        
        # Combinar todos os dataframes
        df_completo = pd.concat(dataframes, ignore_index=True)
        
        # Inicializar variável para contar entregadores cadastrados
        entregadores_cadastrados = []
        
        # Filtrar por entregadores se especificado
        if ids_entregadores:
            df_completo = df_completo[df_completo['id_da_pessoa_entregadora'].isin(ids_entregadores)]
            print(f"📈 Dados após filtro por entregadores: {len(df_completo)} linhas")  # DEBUG
            # Buscar entregadores cadastrados para estatísticas
            entregadores_cadastrados = self._obter_entregadores_cadastrados()
        elif filtrar_por_cadastrados:
            # Buscar lista de entregadores cadastrados no banco
            entregadores_cadastrados = self._obter_entregadores_cadastrados()
            
            print(f"👥 Entregadores cadastrados no banco: {len(entregadores_cadastrados)}")  # DEBUG
            
            if not entregadores_cadastrados:
                raise Exception("Nenhum entregador cadastrado no banco de dados")
            
            # Filtrar apenas os entregadores que estão cadastrados
            df_completo = df_completo[df_completo['id_da_pessoa_entregadora'].isin(entregadores_cadastrados)]
            
            print(f"📈 Dados após filtro por entregadores cadastrados: {len(df_completo)} linhas")  # DEBUG
        else:
            # Não filtrar - processar todos os entregadores do CSV
            print(f"📈 Processando todos os entregadores do CSV: {len(df_completo)} linhas")  # DEBUG
            # Buscar entregadores cadastrados apenas para estatísticas (não para filtrar)
            entregadores_cadastrados = self._obter_entregadores_cadastrados()
        
        if df_completo.empty:
            # Se filtrar_por_cadastrados=True e não houver dados, retornar DataFrame vazio em vez de erro
            if filtrar_por_cadastrados or ids_entregadores:
                print(f"⚠️  Nenhum dado encontrado para os entregadores especificados")
                # Retornar estrutura vazia em vez de erro
                df_vazio = pd.DataFrame(columns=['id_da_pessoa_entregadora', 'recebedor', 'valor_total', 'valor_60_percent', 'valor_final'])
                return {
                    'df_completo': pd.DataFrame(),
                    'consolidado_geral': df_vazio,
                    'total_entregadores': 0,
                    'valor_total_geral': 0.0,
                    'data_processamento': datetime.now().strftime('%d/%m/%Y %H:%M'),
                    'erros': erros,
                    'total_arquivos': len(lista_arquivos),
                    'arquivos_sucesso': len(dataframes),
                    'arquivos_com_erro': len(erros),
                    'total_entregadores_cadastrados': len(entregadores_cadastrados) if entregadores_cadastrados else 0,
                    'entregadores_com_dados': 0
                }
            else:
                raise Exception("Nenhum dado encontrado para os entregadores")
        
        # Consolidar entregadores (passar data_filtro se especificada)
        consolidado_geral = self.consolidar_entregadores(df_completo, data_filtro=data_filtro)
        
        # Calcular estatísticas de forma segura
        total_entregadores = len(consolidado_geral) if consolidado_geral is not None and not consolidado_geral.empty else 0
        valor_total_geral = consolidado_geral['valor_total'].sum() if not consolidado_geral.empty else 0
        
        print(f"🎯 Resultado final: {total_entregadores} entregadores, R$ {valor_total_geral:.2f}")  # DEBUG
        
        return {
            'df_completo': df_completo,
            'consolidado_geral': consolidado_geral,
            'total_entregadores': total_entregadores,
            'valor_total_geral': valor_total_geral,
            'data_processamento': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'erros': erros,
            'total_arquivos': len(lista_arquivos),
            'arquivos_sucesso': len(dataframes),
            'arquivos_com_erro': len(erros),
            'total_entregadores_cadastrados': len(entregadores_cadastrados),
            'entregadores_com_dados': total_entregadores
        }

    def obter_detalhes_processamento_entregador(self, id_entregador, df_completo):
        """Obtém detalhes completos do processamento de um entregador específico"""
        try:
            # Filtrar dados do entregador
            df_entregador = df_completo[df_completo['id_da_pessoa_entregadora'] == id_entregador]
            
            if df_entregador.empty:
                return None
            
            # Consolidar dados específicos
            consolidado = self.consolidar_entregadores(df_entregador)
            
            if consolidado.empty:
                return None
            
            row = consolidado.iloc[0]
            
            # Calcular detalhes
            detalhes = {
                'valor_total': float(row.get('valor_total', 0)),
                'corridas': float(row.get('corridas', 0)),
                'gorjeta': float(row.get('gorjeta', 0)),
                'promo': float(row.get('promo', 0)),
                'online_time': float(row.get('online_time', 0)),
                'rotas_com_ocorrencia': float(row.get('rotas_com_ocorrencia', 0)),
                'tempo_espera': float(row.get('tempo_espera', 0)),
                'outros': float(row.get('outros', 0)),
                'valor_60_percent': float(row.get('valor_60_percent', 0)),
                'valor_final': float(row.get('valor_final', 0)),
                'periodos_trabalhados': list(df_entregador['periodo'].unique()),
                'subpracas': list(df_entregador['subpraca'].unique())
            }
            
            return detalhes
            
        except Exception as e:
            print(f"Erro ao obter detalhes do processamento: {str(e)}")
            return None
    
    def gerar_relatorio_excel(self, consolidado_geral, caminho_saida):
        """Gera relatório completo em Excel"""
        try:
            with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
                # Worksheet consolidado
                consolidado_geral.to_excel(writer, sheet_name='Consolidado', index=False)
                
                # Worksheet resumo por praça
                resumo_praca = consolidado_geral.groupby('pracas').agg({
                    'valor_total': 'sum',
                    'valor_60_percent': 'sum',
                    'corridas': 'sum',
                    'gorjeta': 'sum'
                }).round(2)
                resumo_praca.to_excel(writer, sheet_name='Resumo por Praça')
                
                # Worksheet para pagamento
                pagamento_df = consolidado_geral[['recebedor', 'id_da_pessoa_entregadora', 'valor_final', 'subpracas']].copy()
                pagamento_df.columns = ['Entregador', 'ID', 'Valor a Pagar', 'Sub-Praças']
                pagamento_df.to_excel(writer, sheet_name='Para Pagamento', index=False)
            
            # Formatar Excel
            self._formatar_excel(caminho_saida)
            
            return caminho_saida
            
        except Exception as e:
            raise Exception(f"Erro ao gerar relatório Excel: {str(e)}")
    
    def _formatar_excel(self, caminho_arquivo):
        """Aplica formatação ao arquivo Excel gerado"""
        wb = load_workbook(caminho_arquivo)
        azul_abjp = "0B5CFF"
        cinza_claro = "F2F2F2"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Formatar cabeçalho
            for cell in ws[1]:
                cell.fill = PatternFill(start_color=azul_abjp, end_color=azul_abjp, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Formatar células de dados
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    # Formato monetário para colunas de valor
                    if cell.column_letter in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = 'R$ #,##0.00'
                    
                    # Alternar cores das linhas
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill(start_color=cinza_claro, end_color=cinza_claro, fill_type="solid")

        wb.save(caminho_arquivo)
    
    # Método de debug para verificar todos os métodos disponíveis
    def listar_metodos(self):
        """Lista todos os métodos disponíveis na classe"""
        return [method for method in dir(self) if not method.startswith('_')]