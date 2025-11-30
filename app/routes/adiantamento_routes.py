"""
Rotas para gerenciamento de adiantamentos
"""
import os
import json
import csv
from datetime import datetime, date
import pandas as pd
from flask import render_template, request, redirect, url_for, flash, send_file, jsonify, make_response
from app.utils.auth_decorators import login_required, master_required
from app.models.database import get_db_connection
from app.utils.path_manager import get_week_folder
from config import Config
from app.services.processador_csv_service import ProcessadorCSVService
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.utils.form_control import (
    get_form_config,
    abrir_formulario,
    fechar_formulario,
    agendar_abertura,
    agendar_fechamento,
    form_is_open
)
from app.utils.constants import (
    TEMPLATES_ADIANTAMENTO,
    TEMPLATES_ADMIN,
    MESSAGES,
    ARQUIVO_SOLICITACOES,
    ARQUIVO_ULTIMO_CONSOLIDADO,
    ARQUIVO_CONSOLIDADO_DIARIO,
    FORMATO_DATA_SQL,
    FORMATO_DATA_ISO,
    FORMATO_DATA_DATETIME_LOCAL,
    FORMATO_DATA_DIARIO,
    FORMATO_DATA_ARQUIVO,
    ACOES_LOG
)
from app.utils.route_helpers import (
    normalize_cpf,
    format_datetime_local_to_sql,
    is_ajax_request,
    json_response,
    get_flash_message
)


def _salvar_resposta_json(resposta):
    """Salva as respostas no banco de dados (JSON removido - usando apenas PostgreSQL)"""
    # As respostas já são salvas na tabela solicitacoes_adiantamento
    # Esta função agora é apenas um placeholder para compatibilidade
    # O salvamento real acontece em criar_solicitacao_adiantamento()
    pass


def _aplicar_filtros_solicitacoes(solicitacoes, busca, filtro_dia, filtro_mes, filtro_cpf_status, filtro_sub):
    """Aplica filtros em memória às solicitações"""
    if busca:
        termo = busca.lower()
        solicitacoes = [
            s for s in solicitacoes
            if termo in (s.get('nome') or '').lower()
            or termo in (s.get('email') or '').lower()
            or termo in (s.get('cpf') or '').lower()
        ]
    
    if filtro_dia:
        solicitacoes = [
            s for s in solicitacoes
            if (s.get('data_envio') or '')[:10] == filtro_dia
        ]
    
    if filtro_mes:
        solicitacoes = [
            s for s in solicitacoes
            if (s.get('data_envio') or '')[5:7] == f"{int(filtro_mes):02d}"
        ]
    
    if filtro_cpf_status in ('0', '1'):
        flag = int(filtro_cpf_status)
        solicitacoes = [
            s for s in solicitacoes
            if int(s.get('cpf_bate', 0)) == flag
        ]
    
    if filtro_sub:
        solicitacoes = [
            s for s in solicitacoes
            if (s.get('praca') or '') == filtro_sub
        ]
    
    return solicitacoes


def _calcular_valores_dia_solicitacoes(solicitacoes, data_ref):
    """
    Calcula os valores do dia específico para cada solicitação de adiantamento
    IMPORTANTE: data_ref deve ser a data do período de referência (quando o entregador trabalhou),
    não a data de envio da solicitação!
    
    OTIMIZAÇÃO: Processa apenas o ÚLTIMO arquivo CSV enviado para melhor performance
    """
    from app.services.processador_csv_service import ProcessadorCSVService
    from app.utils.path_manager import get_week_folder
    from config import Config
    import pandas as pd
    
    # Coletar IDs dos entregadores que solicitaram
    ids_entregadores = set()
    
    for s in solicitacoes:
        if s.get('id_da_pessoa_entregadora'):
            ids_entregadores.add(str(s['id_da_pessoa_entregadora']))
    
    if not ids_entregadores:
        return solicitacoes
    
    # Buscar apenas o ÚLTIMO arquivo CSV (mais recente)
    pasta_uploads = get_week_folder(Config.UPLOAD_FOLDER)
    todos_arquivos = [
        os.path.join(pasta_uploads, f)
        for f in os.listdir(pasta_uploads)
        if f.endswith('.csv') and f != ARQUIVO_ULTIMO_CONSOLIDADO
    ]
    
    if not todos_arquivos:
        return solicitacoes
    
    # Pegar apenas o último arquivo (mais recente por data de modificação)
    ultimo_arquivo = max(todos_arquivos, key=os.path.getmtime)
    nome_ultimo = os.path.basename(ultimo_arquivo)
    
    print(f"📄 Calculando valores usando apenas o último arquivo: {nome_ultimo}")
    
    try:
        # Processar apenas o último arquivo CSV, filtrando pela data do período de referência
        # e apenas para os entregadores que solicitaram
        processador = ProcessadorCSVService()
        resultado = processador.processar_multiplos_csv(
            [ultimo_arquivo],  # Apenas o último arquivo
            data_filtro=data_ref,  # Esta data será usada para filtrar dentro do CSV
            ids_entregadores=list(ids_entregadores)  # Filtrar apenas entregadores que solicitaram
        )
        
        df_consolidado = resultado['consolidado_geral']
        
        if df_consolidado is None or df_consolidado.empty:
            print(f"⚠️  Nenhum dado encontrado para a data {data_ref} e entregadores {ids_entregadores}")
            return solicitacoes
        
        # Criar dicionário de valores por ID do entregador
        valores_por_id = {}
        for _, row in df_consolidado.iterrows():
            id_ent = str(row.get('id_da_pessoa_entregadora', ''))
            if id_ent:
                valores_por_id[id_ent] = {
                    'valor_total': float(row.get('valor_total', 0)),
                    'valor_60_percent': float(row.get('valor_60_percent', 0)),
                    'valor_final': float(row.get('valor_final', 0)),
                    'gorjeta': float(row.get('gorjeta', 0)),
                    'corridas': float(row.get('corridas', 0)),
                    'promo': float(row.get('promo', 0)),
                    'online_time': float(row.get('online_time', 0))
                }
                print(f"✅ Valores encontrados para entregador {id_ent}: Total={valores_por_id[id_ent]['valor_total']}, Final={valores_por_id[id_ent]['valor_final']}")
        
        # Atualizar solicitações com valores calculados do dia
        for s in solicitacoes:
            id_ent = s.get('id_da_pessoa_entregadora')
            if id_ent and str(id_ent) in valores_por_id:
                valores = valores_por_id[str(id_ent)]
                s['valor_total_dia'] = valores['valor_total']
                s['valor_60_percent_dia'] = valores['valor_60_percent']
                s['valor_final_dia'] = valores['valor_final']
                s['gorjeta_dia'] = valores['gorjeta']
                s['corridas_dia'] = valores['corridas']
                s['promo_dia'] = valores['promo']
                s['online_time_dia'] = valores['online_time']
            else:
                # Se não encontrou valores, definir como 0
                s['valor_total_dia'] = 0.0
                s['valor_60_percent_dia'] = 0.0
                s['valor_final_dia'] = 0.0
                s['gorjeta_dia'] = 0.0
                s['corridas_dia'] = 0.0
                s['promo_dia'] = 0.0
                s['online_time_dia'] = 0.0
                print(f"⚠️  Nenhum valor encontrado para entregador {id_ent} na data {data_ref}")
        
    except Exception as e:
        print(f"❌ Erro ao calcular valores do dia: {str(e)}")
        import traceback
        traceback.print_exc()
        # Em caso de erro, retornar solicitações sem valores calculados
    
    return solicitacoes


def init_adiantamento_routes(app):
    """Inicializa as rotas de adiantamento"""
    
    @app.route('/adiantamento/admin', methods=['GET'])
    @login_required
    def lista_solicitacoes():
        """Lista todas as solicitações de adiantamento, com filtros"""
        busca = (request.args.get('busca') or '').strip()
        # Se não houver filtro de dia, usar o dia atual como padrão
        filtro_dia = (request.args.get('dia') or '').strip()
        if not filtro_dia:
            filtro_dia = date.today().strftime(FORMATO_DATA_ISO)
        filtro_mes = (request.args.get('mes') or '').strip()
        filtro_cpf_status = (request.args.get('cpf_status') or '').strip()
        filtro_sub = (request.args.get('sub') or '').strip()
        
        conn = get_db_connection()
        from app.models.database import is_postgresql_connection
        is_postgresql = is_postgresql_connection(conn)
        
        if is_postgresql:
            from psycopg2.extras import RealDictCursor
            cursor = conn.cursor(cursor_factory=RealDictCursor)
        else:
            cursor = conn.cursor()
        
        # Buscar arquivos CSV enviados (com nome e data de upload)
        pasta_uploads = get_week_folder(Config.UPLOAD_FOLDER)
        arquivos_csv_info = []
        
        if os.path.exists(pasta_uploads):
            for f in os.listdir(pasta_uploads):
                if f.endswith('.csv') and f != ARQUIVO_ULTIMO_CONSOLIDADO:
                    caminho_completo = os.path.join(pasta_uploads, f)
                    # Extrair data de upload do nome do arquivo (formato: YYYYMMDD_HHMMSS_nome.csv)
                    try:
                        partes = f.replace('.csv', '').split('_')
                        if len(partes) >= 2:
                            data_upload_str = partes[0]  # YYYYMMDD
                            hora_upload_str = partes[1]  # HHMMSS
                            # Converter para datetime
                            data_upload = datetime.strptime(f"{data_upload_str}_{hora_upload_str}", '%Y%m%d_%H%M%S')
                            # Nome original do arquivo (tudo depois da segunda parte)
                            nome_original = '_'.join(partes[2:]) if len(partes) > 2 else f
                            
                            arquivos_csv_info.append({
                                'nome_arquivo': f,
                                'nome_original': nome_original,
                                'data_upload': data_upload.date(),
                                'hora_upload': data_upload.time(),
                                'data_upload_formatada': data_upload.strftime('%d/%m/%Y %H:%M'),
                                'caminho': caminho_completo
                            })
                    except (ValueError, IndexError):
                        # Se não conseguir extrair data, usar data de modificação do arquivo
                        data_mod = datetime.fromtimestamp(os.path.getmtime(caminho_completo))
                        arquivos_csv_info.append({
                            'nome_arquivo': f,
                            'nome_original': f,
                            'data_upload': data_mod.date(),
                            'hora_upload': data_mod.time(),
                            'data_upload_formatada': data_mod.strftime('%d/%m/%Y %H:%M'),
                            'caminho': caminho_completo
                        })
        
        # Ordenar por data de upload (mais recente primeiro)
        arquivos_csv_info.sort(key=lambda x: (x['data_upload'], x['hora_upload']), reverse=True)
        
        # Manter compatibilidade: também buscar dias disponíveis das solicitações
        # Verificar se a coluna data_envio existe
        try:
            if is_postgresql:
                cursor.execute("""
                    SELECT DISTINCT DATE(data_envio) as dia
                    FROM solicitacoes_adiantamento
                    WHERE data_envio IS NOT NULL
                    ORDER BY dia DESC
                """)
            else:
                cursor.execute("""
                    SELECT DISTINCT DATE(data_envio) as dia
                    FROM solicitacoes_adiantamento
                    ORDER BY dia DESC
                """)
            dias_disponiveis = [r["dia"] for r in cursor.fetchall()]
        except Exception:
            # Se a coluna não existir, usar lista vazia
            dias_disponiveis = []
        
        # Buscar todas as solicitações
        # Normalização mais robusta de CPF: remove todos os caracteres não numéricos
        try:
            if is_postgresql:
                cursor.execute("""
                    SELECT 
                        s.id, s.email, s.nome, s.cpf,
                        CASE 
                            WHEN e.id_da_pessoa_entregadora IS NOT NULL THEN 1 ELSE 0
                        END AS cpf_bate,
                        s.praca, s.valor_informado, s.concorda, s.data_envio,
                        e.recebedor AS recebedor_base,
                        e.id_da_pessoa_entregadora
                    FROM solicitacoes_adiantamento s
                    LEFT JOIN entregadores e 
                        ON REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(s.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '') =
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(e.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '')
                    ORDER BY s.data_envio DESC NULLS LAST
                """)
            else:
                cursor.execute("""
                    SELECT 
                        s.id, s.email, s.nome, s.cpf,
                        CASE 
                            WHEN e.id_da_pessoa_entregadora IS NOT NULL THEN 1 ELSE 0
                        END AS cpf_bate,
                        s.praca, s.valor_informado, s.concorda, s.data_envio,
                        e.recebedor AS recebedor_base,
                        e.id_da_pessoa_entregadora
                    FROM solicitacoes_adiantamento s
                    LEFT JOIN entregadores e 
                        ON REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(s.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '') =
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(e.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '')
                    ORDER BY s.data_envio DESC
                """)
            solicitacoes = [dict(r) for r in cursor.fetchall()]
        except Exception as e:
            # Se houver erro (coluna não existe), buscar sem data_envio
            print(f"⚠️ Aviso ao buscar solicitações: {e}")
            solicitacoes = []
        
        conn.close()
        
        subpracas = sorted({s.get('praca') for s in solicitacoes if s.get('praca')})
        
        # Aplicar filtros
        solicitacoes = _aplicar_filtros_solicitacoes(
            solicitacoes, busca, filtro_dia, filtro_mes, filtro_cpf_status, filtro_sub
        )
        
        # Calcular valores do dia para cada solicitação APENAS se houver filtro de dia
        # Se não houver filtro, não calcular (evita processar muitos arquivos desnecessariamente)
        if filtro_dia:
            try:
                data_para_calculo = datetime.strptime(filtro_dia, FORMATO_DATA_ISO).date()
                solicitacoes = _calcular_valores_dia_solicitacoes(solicitacoes, data_para_calculo)
            except (ValueError, Exception) as e:
                print(f"Erro ao calcular valores do dia: {str(e)}")
        
        return render_template(
            TEMPLATES_ADIANTAMENTO['lista'],
            solicitacoes=solicitacoes,
            subpracas=subpracas,
            dias_disponiveis=dias_disponiveis,
            arquivos_csv_info=arquivos_csv_info,  # Adicionar informações dos arquivos CSV
            filtro_busca=busca,
            filtro_dia=filtro_dia,
            filtro_mes=filtro_mes,
            filtro_cpf_status=filtro_cpf_status,
            filtro_sub=filtro_sub,
            current_date=date.today().isoformat()
        )
    
    @app.route('/adiantamento', methods=['GET'])
    def formulario():
        """Exibe o formulário de adiantamento somente se estiver aberto"""
        if not form_is_open():
            return render_template(TEMPLATES_ADIANTAMENTO['fechado'])
        return render_template(TEMPLATES_ADIANTAMENTO['form_public'])
    
    @app.route('/adiantamento/enviar', methods=['POST'])
    def enviar():
        """Recebe o formulário de adiantamento e salva JSON + banco"""
        email = request.form.get('email')
        nome = request.form.get('nome')
        cpf = request.form.get('cpf')
        praca = request.form.get('praca')
        valor = request.form.get('valor')
        concorda = request.form.get('concorda')
        
        if not cpf or not email:
            return redirect(url_for('formulario'))
        
        # Normalizar CPF para busca
        cpf_limpo = normalize_cpf(cpf)
        email_limpo = email.strip().lower()
        
        conn = get_db_connection()
        from app.models.database import is_postgresql_connection
        is_postgresql = is_postgresql_connection(conn)
        placeholder = "%s" if is_postgresql else "?"
        
        if is_postgresql:
            from psycopg2.extras import RealDictCursor
            cursor = conn.cursor(cursor_factory=RealDictCursor)
        else:
            cursor = conn.cursor()
        
        # Validar se o entregador está cadastrado com CPF e email
        # Buscar todos os entregadores e normalizar CPF em Python
        # No PostgreSQL, usar LENGTH para strings vazias
        if is_postgresql:
            cursor.execute("SELECT id_da_pessoa_entregadora, recebedor, cpf, email FROM entregadores WHERE cpf IS NOT NULL AND LENGTH(TRIM(cpf)) > 0")
        else:
            cursor.execute("SELECT id_da_pessoa_entregadora, recebedor, cpf, email FROM entregadores WHERE cpf IS NOT NULL AND cpf != ''")
        todos_entregadores = cursor.fetchall()
        
        entregador = None
        for e in todos_entregadores:
            if isinstance(e, dict):
                cpf_entregador_limpo = normalize_cpf(e.get('cpf') or '')
            else:
                cpf_entregador_limpo = normalize_cpf(e[2] or '')  # cpf está na posição 2
            if cpf_entregador_limpo == cpf_limpo:
                entregador = e
                break
        
        if not entregador:
            conn.close()
            return render_template(
                TEMPLATES_ADIANTAMENTO['form_public'],
                erro="CPF não encontrado no sistema. Entre em contato com o suporte."
            )
        
        # Verificar se o email informado corresponde ao email cadastrado
        if isinstance(entregador, dict):
            email_cadastrado = (entregador.get('email') or '').strip().lower()
        else:
            email_cadastrado = (entregador[3] or '').strip().lower()  # email está na posição 3
        
        if email_cadastrado and email_cadastrado != email_limpo:
            conn.close()
            return render_template(
                TEMPLATES_ADIANTAMENTO['form_public'],
                erro=f"Email não corresponde ao cadastrado."
            )
        
        # Verificar se o entregador tem email cadastrado
        if not email_cadastrado:
            conn.close()
            return render_template(
                TEMPLATES_ADIANTAMENTO['form_public'],
                erro="Email não cadastrado no sistema. Entre em contato com o suporte para cadastrar seu email."
            )
        
        # Regra: apenas 1 solicitação por CPF por dia
        hoje = date.today().strftime(FORMATO_DATA_ISO)
        try:
            cursor.execute(f"""
                SELECT COUNT(*) FROM solicitacoes_adiantamento
                WHERE cpf = {placeholder} AND DATE(data_envio) = {placeholder}
            """, (cpf, hoje))
            resultado = cursor.fetchone()
            ja_existe = resultado[0] if resultado else 0
        except Exception:
            # Se a coluna data_envio não existir, permitir múltiplas solicitações
            ja_existe = 0
        
        if ja_existe > 0:
            conn.close()
            return render_template(
                TEMPLATES_ADIANTAMENTO['bloqueado'],
                nome=nome,
                cpf=cpf
            )
        
        # Preparar dados JSON completos
        resposta_json = {
            "email": email,
            "nome": nome,
            "cpf": cpf,
            "praca": praca,
            "valor": valor,
            "concorda": concorda,
            "data_envio": datetime.now().strftime(FORMATO_DATA_SQL)
        }
        
        # Salvar no banco (com JSON completo)
        data_envio = resposta_json["data_envio"]
        dados_json_str = json.dumps(resposta_json, ensure_ascii=False)
        
        # Verificar se é PostgreSQL ou SQLite
        from app.models.database import is_postgresql_connection
        is_postgresql = is_postgresql_connection(conn)
        
        try:
            if is_postgresql:
                cursor.execute("""
                    INSERT INTO solicitacoes_adiantamento 
                    (email, nome, cpf, praca, valor_informado, concorda, data_envio, dados_json)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                """, (email, nome, cpf, praca, valor, concorda, data_envio, dados_json_str))
            else:
                cursor.execute("""
                    INSERT INTO solicitacoes_adiantamento 
                    (email, nome, cpf, praca, valor_informado, concorda, data_envio, dados_json)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (email, nome, cpf, praca, valor, concorda, data_envio, dados_json_str))
        except Exception as e:
            # Se falhar (ex: coluna dados_json não existe), tentar sem dados_json
            print(f"Erro ao inserir com dados_json, tentando sem: {e}")
            if is_postgresql:
                cursor.execute("""
                    INSERT INTO solicitacoes_adiantamento 
                    (email, nome, cpf, praca, valor_informado, concorda, data_envio)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (email, nome, cpf, praca, valor, concorda, data_envio))
            else:
                cursor.execute("""
                    INSERT INTO solicitacoes_adiantamento 
                    (email, nome, cpf, praca, valor_informado, concorda, data_envio)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (email, nome, cpf, praca, valor, concorda, data_envio))
        
        conn.commit()
        conn.close()
        
        # Dados já salvos no banco (tabela solicitacoes_adiantamento)
        # JSON removido - usando apenas PostgreSQL para segurança e consultas
        
        return render_template(TEMPLATES_ADIANTAMENTO['sucesso'], nome=nome)
    
    @app.route('/adiantamento/gerar-diario', methods=['GET'])
    def gerar_diario_adiantamento():
        """Gera um CSV diário com quem solicitou adiantamento"""
        # Verificar se foi passado um arquivo específico ou uma data
        arquivo_nome = request.args.get('arquivo')
        data_str = request.args.get('data')
        
        # Se foi passado um arquivo, usar esse arquivo específico
        arquivo_especifico = None
        if arquivo_nome:
            pasta_uploads = get_week_folder(Config.UPLOAD_FOLDER)
            caminho_arquivo = os.path.join(pasta_uploads, arquivo_nome)
            
            if not os.path.exists(caminho_arquivo):
                flash(f"Arquivo {arquivo_nome} não encontrado.", "adiantamento_error")
                return redirect(url_for('lista_solicitacoes'))
            
            arquivo_especifico = caminho_arquivo
            # Usar data de hoje para buscar solicitações
            data_ref = date.today()
        elif data_str:
            try:
                data_ref = datetime.strptime(data_str, FORMATO_DATA_ISO).date()
            except ValueError:
                flash(
                    get_flash_message('adiantamento', 'data_invalida'),
                    "adiantamento_error"
                )
                return redirect(url_for('lista_solicitacoes'))
        else:
            data_ref = date.today()
        
        conn = get_db_connection()
        from app.models.database import is_postgresql_connection
        is_postgresql = is_postgresql_connection(conn)
        placeholder = "%s" if is_postgresql else "?"
        
        if is_postgresql:
            from psycopg2.extras import RealDictCursor
            cursor = conn.cursor(cursor_factory=RealDictCursor)
        else:
            cursor = conn.cursor()
        
        # Se data_ref for None (arquivo específico), buscar TODAS as solicitações
        if data_ref is None:
            try:
                cursor.execute("""
                    SELECT s.cpf, s.nome, s.praca, s.valor_informado, s.concorda, s.data_envio,
                           e.id_da_pessoa_entregadora, e.recebedor AS recebedor_base
                    FROM solicitacoes_adiantamento s
                    LEFT JOIN entregadores e 
                        ON REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(s.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '') =
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(e.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '')
                    ORDER BY s.data_envio ASC NULLS LAST
                """)
            except Exception:
                # Se data_envio não existir, buscar sem ordenar por data_envio
                cursor.execute("""
                    SELECT s.cpf, s.nome, s.praca, s.valor_informado, s.concorda,
                           e.id_da_pessoa_entregadora, e.recebedor AS recebedor_base
                    FROM solicitacoes_adiantamento s
                    LEFT JOIN entregadores e 
                        ON REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(s.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '') =
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(e.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '')
                """)
        else:
            try:
                cursor.execute(f"""
                    SELECT s.cpf, s.nome, s.praca, s.valor_informado, s.concorda, s.data_envio,
                           e.id_da_pessoa_entregadora, e.recebedor AS recebedor_base
                    FROM solicitacoes_adiantamento s
                    LEFT JOIN entregadores e 
                        ON REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(s.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '') =
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(e.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '')
                    WHERE DATE(s.data_envio) = {placeholder}
                    ORDER BY s.data_envio ASC
                """, (data_ref.strftime(FORMATO_DATA_ISO),))
            except Exception:
                # Se data_envio não existir, buscar todas sem filtro de data
                cursor.execute("""
                    SELECT s.cpf, s.nome, s.praca, s.valor_informado, s.concorda,
                           e.id_da_pessoa_entregadora, e.recebedor AS recebedor_base
                    FROM solicitacoes_adiantamento s
                    LEFT JOIN entregadores e 
                        ON REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(s.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '') =
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                            LTRIM(RTRIM(COALESCE(e.cpf, ''))), 
                            '.', ''), '-', ''), ' ', ''), '(', ''), ')', ''), '/', '')
                """)
        
        linhas = cursor.fetchall()
        conn.close()
        
        if not linhas:
            flash(
                get_flash_message('adiantamento', 'nenhuma_solicitacao', 
                                date=data_ref.strftime('%d/%m/%Y')),
                "adiantamento_info"
            )
            return redirect(url_for('lista_solicitacoes'))
        
        # Montar listas de CPFs e IDs
        cpfs_solicitantes = set()
        ids_solicitantes = set()
        
        for row in linhas:
            cpf_raw = row['cpf'] if isinstance(row, dict) else row[0]
            id_ent = row['id_da_pessoa_entregadora'] if isinstance(row, dict) else row[6]
            cpf_norm = normalize_cpf(cpf_raw)
            if cpf_norm:
                cpfs_solicitantes.add(cpf_norm)
            if id_ent:
                ids_solicitantes.add(str(id_ent))
        
        # Determinar qual arquivo usar
        if arquivo_especifico:
            # Usar o arquivo específico passado como parâmetro
            arquivo_para_processar = arquivo_especifico
            nome_arquivo = os.path.basename(arquivo_para_processar)
            print(f"📄 Gerando relatório usando arquivo específico: {nome_arquivo}")
        else:
            # Usar o último arquivo enviado
            pasta_uploads = get_week_folder(Config.UPLOAD_FOLDER)
            todos_arquivos = [
                os.path.join(pasta_uploads, f)
                for f in os.listdir(pasta_uploads)
                if f.endswith('.csv') and f != ARQUIVO_ULTIMO_CONSOLIDADO
            ]
            
            if not todos_arquivos:
                flash("Nenhum CSV encontrado para processar.", "adiantamento_error")
                return redirect(url_for('lista_solicitacoes'))
            
            arquivo_para_processar = max(todos_arquivos, key=os.path.getmtime)
            nome_arquivo = os.path.basename(arquivo_para_processar)
            print(f"📄 Gerando relatório usando o último arquivo: {nome_arquivo}")
        
        # Processar o arquivo CSV (sem filtrar por data do período, processar tudo)
        processador = ProcessadorCSVService()
        try:
            resultado = processador.processar_multiplos_csv(
                [arquivo_para_processar],  # Arquivo específico ou último
                data_filtro=None,  # Não filtrar por data do período, processar tudo do arquivo
                ids_entregadores=list(ids_solicitantes) if ids_solicitantes else None
            )
            df_consolidado = resultado['consolidado_geral']
        except Exception as e:
            flash(
                f"Erro ao processar arquivo: {str(e)}",
                "adiantamento_error"
            )
            return redirect(url_for('lista_solicitacoes'))
        
        if df_consolidado is None or df_consolidado.empty:
            flash(
                f"Nenhum dado encontrado no arquivo {nome_arquivo} para os entregadores solicitantes.",
                "adiantamento_warning"
            )
            return redirect(url_for('lista_solicitacoes'))
        
        # Normalizar CPF no consolidado
        if 'cpf' in df_consolidado.columns:
            df_consolidado['cpf_norm'] = df_consolidado['cpf'].apply(normalize_cpf)
        else:
            df_consolidado['cpf_norm'] = ''
        
        if 'id_da_pessoa_entregadora' in df_consolidado.columns:
            df_consolidado['id_da_pessoa_entregadora'] = df_consolidado['id_da_pessoa_entregadora'].astype(str)
        else:
            df_consolidado['id_da_pessoa_entregadora'] = ''
        
        # Filtrar apenas quem pediu adiantamento (por CPF ou ID)
        mask = (
            df_consolidado['cpf_norm'].isin(cpfs_solicitantes) |
            df_consolidado['id_da_pessoa_entregadora'].isin(ids_solicitantes)
        )
        df_diario = df_consolidado[mask].copy()
        
        if df_diario.empty:
            flash(
                get_flash_message('adiantamento', 'nenhum_registro'),
                "warning"
            )
            return redirect(url_for('lista_solicitacoes'))
        
        # Salvar XLSX
        pasta_relatorios = get_week_folder(Config.RELATORIOS_FOLDER)
        os.makedirs(pasta_relatorios, exist_ok=True)
        
        nome_arquivo = f"diario_adiantamento_{data_ref.strftime(FORMATO_DATA_DIARIO)}.xlsx"
        caminho_saida = os.path.join(pasta_relatorios, nome_arquivo)
        
        df_diario.to_excel(caminho_saida, index=False, engine='openpyxl')
        
        flash(
            get_flash_message('adiantamento', 'diario_gerado', filename=nome_arquivo),
            "success"
        )
        return send_file(caminho_saida, as_attachment=True)
    
    @app.route("/adiantamento/admin/forms", methods=["GET"])
    @login_required
    def admin_forms_painel():
        cfg = get_form_config()
        if cfg.get("days_enabled") is None:
            cfg["days_enabled"] = ""
        
        # Sempre retornar a página completa (não mais modal)
        return render_template(TEMPLATES_ADMIN['form_config'], cfg=cfg)
    
    @app.route("/adiantamento/admin/forms/config", methods=["GET"])
    def admin_forms_config_reload():
        """Retorna somente o conteúdo do modal (AJAX)"""
        cfg = get_form_config()
        if cfg.get("days_enabled") is None:
            cfg["days_enabled"] = ""
        return render_template(TEMPLATES_ADMIN['form_config_modal'], cfg=cfg)
    
    @app.route("/adiantamento/admin/forms/open", methods=["POST"])
    def admin_forms_open():
        abrir_formulario()
        flash(
            get_flash_message('adiantamento', 'form_aberto'),
            "success"
        )
        if is_ajax_request():
            return json_response(
                success=True,
                message=get_flash_message('adiantamento', 'form_aberto'),
                category="adiantamento_success"
            )
        return redirect(url_for("admin_forms_painel"))
    
    @app.route("/adiantamento/admin/forms/close", methods=["POST"])
    def admin_forms_close():
        fechar_formulario()
        flash(
            get_flash_message('adiantamento', 'form_fechado'),
            "adiantamento_info"
        )
        if is_ajax_request():
            return json_response(
                success=True,
                message=get_flash_message('adiantamento', 'form_fechado'),
                category="adiantamento_info"
            )
        return redirect(url_for("admin_forms_painel"))
    
    @app.route("/adiantamento/admin/forms/schedule", methods=["POST"])
    def admin_forms_schedule():
        abrir = request.form.get("scheduled_open")
        fechar = request.form.get("scheduled_close")
        
        abrir = format_datetime_local_to_sql(abrir) if abrir else None
        fechar = format_datetime_local_to_sql(fechar) if fechar else None
        
        agendar_abertura(abrir)
        agendar_fechamento(fechar)
        
        flash(
            get_flash_message('adiantamento', 'agendamento_atualizado'),
            "adiantamento_success"
        )
        if is_ajax_request():
            return json_response(
                success=True,
                message=get_flash_message('adiantamento', 'agendamento_atualizado'),
                category="adiantamento_success"
            )
        return redirect(url_for("admin_forms_painel"))
    
    @app.route("/adiantamento/admin/logs", methods=["GET"])
    @master_required
    def admin_form_logs():
        """Exibe a página com todos os logs do formulário"""
        acao_filtro = request.args.get("acao", "")
        data_inicio = request.args.get("inicio", "")
        data_fim = request.args.get("fim", "")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = "SELECT * FROM form_logs WHERE 1=1"
        params = []
        
        if acao_filtro:
            query += " AND acao = ?"
            params.append(acao_filtro)
        
        if data_inicio:
            query += " AND date(data_hora) >= date(?)"
            params.append(data_inicio)
        
        if data_fim:
            query += " AND date(data_hora) <= date(?)"
            params.append(data_fim)
        
        query += " ORDER BY id DESC"
        cursor.execute(query, params)
        logs = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        acoes_unicas = list(ACOES_LOG.values())
        
        return render_template(
            TEMPLATES_ADMIN['form_logs'],
            logs=logs,
            acoes=acoes_unicas,
            acao_filtro=acao_filtro,
            data_inicio=data_inicio,
            data_fim=data_fim
        )
    
    @app.route("/adiantamento/admin/logs/exportar", methods=["GET"])
    @master_required
    def admin_form_logs_exportar():
        """Exporta os logs em CSV"""
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM form_logs ORDER BY id DESC")
        rows = cursor.fetchall()
        conn.close()
        
        si = []
        si.append("id,acao,detalhe,data_hora\n")
        for r in rows:
            linha = f'{r["id"]},{r["acao"]},"{r["detalhe"]}",{r["data_hora"]}\n'
            si.append(linha)
        
        resposta = make_response("".join(si))
        resposta.headers["Content-Disposition"] = "attachment; filename=form_logs.csv"
        resposta.headers["Content-Type"] = "text/csv"
        return resposta
    
    @app.route("/adiantamento/admin/forms/auto", methods=["POST"])
    def admin_forms_auto():
        modo = int(request.form.get("auto_mode", 0))
        abre = request.form.get("auto_open_time")
        fecha = request.form.get("auto_close_time")
        dias = request.form.getlist("days_enabled")
        dias_str = ",".join(dias) if dias else None
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE form_config
            SET auto_mode = ?, auto_open_time = ?, auto_close_time = ?, days_enabled = ?
            WHERE id = 1
        """, (modo, abre, fecha, dias_str))
        conn.commit()
        conn.close()
        
        flash(
            get_flash_message('adiantamento', 'config_salva'),
            "adiantamento_success"
        )
        if is_ajax_request():
            return json_response(
                success=True,
                message=get_flash_message('adiantamento', 'config_salva'),
                category="adiantamento_success"
            )
        return redirect(url_for("admin_forms_painel"))
    
    @app.route('/adiantamento/admin/exportar-excel', methods=['GET'])
    def exportar_excel_adiantamento():
        """Exporta consolidado normal e consolidado do forms em Excel"""
        try:
            tipo = request.args.get('tipo', 'normal')  # 'normal' ou 'forms'
            
            pasta_relatorios = get_week_folder(Config.RELATORIOS_FOLDER)
            os.makedirs(pasta_relatorios, exist_ok=True)
            
            pasta_uploads = get_week_folder(Config.UPLOAD_FOLDER)
            
            # Carregar consolidado
            if tipo == 'forms':
                consolidado_path = os.path.join(pasta_uploads, ARQUIVO_CONSOLIDADO_DIARIO)
                nome_arquivo = f"consolidado_forms_{datetime.now().strftime(FORMATO_DATA_ARQUIVO)}.xlsx"
            else:
                consolidado_path = os.path.join(pasta_uploads, ARQUIVO_ULTIMO_CONSOLIDADO)
                nome_arquivo = f"consolidado_normal_{datetime.now().strftime(FORMATO_DATA_ARQUIVO)}.xlsx"
            
            if not os.path.exists(consolidado_path):
                flash('Nenhum consolidado encontrado. Processe um CSV primeiro.', 'adiantamento_error')
                return redirect(url_for('lista_solicitacoes'))
            
            # Ler consolidado CSV
            df_consolidado = pd.read_csv(consolidado_path, encoding='utf-8')
            
            # Buscar dados dos entregadores do banco
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Buscar todos os entregadores com suas informações
            # Para chave PIX, buscar a mais recente de cada entregador
            cursor.execute("""
                SELECT 
                    e.id_da_pessoa_entregadora,
                    e.recebedor,
                    e.email,
                    e.cnpj,
                    e.emissor,
                    e.subpraca,
                    (SELECT h2.chave_pix 
                     FROM historico_pix h2 
                     WHERE h2.id_da_pessoa_entregadora = e.id_da_pessoa_entregadora 
                     ORDER BY h2.data_registro DESC 
                     LIMIT 1) as chave_pix,
                    (SELECT h2.tipo_de_chave_pix 
                     FROM historico_pix h2 
                     WHERE h2.id_da_pessoa_entregadora = e.id_da_pessoa_entregadora 
                     ORDER BY h2.data_registro DESC 
                     LIMIT 1) as tipo_de_chave_pix
                FROM entregadores e
            """)
            
            entregadores_data = cursor.fetchall()
            conn.close()
            
            # Criar DataFrame de entregadores
            df_entregadores = pd.DataFrame([dict(row) for row in entregadores_data])
            
            # Renomear colunas do banco para evitar conflito no merge
            df_entregadores = df_entregadores.rename(columns={
                'recebedor': 'recebedor_db',
                'email': 'email_db',
                'cnpj': 'cnpj_db',
                'emissor': 'emissor_db',
                'subpraca': 'subpraca_db',
                'chave_pix': 'chave_pix_db',
                'tipo_de_chave_pix': 'tipo_de_chave_pix_db'
            })
            
            # Fazer merge com consolidado
            df_final = pd.merge(
                df_consolidado,
                df_entregadores,
                on='id_da_pessoa_entregadora',
                how='left'
            )
            
            # Calcular adiantamento como 60% do valor_total (excluindo gorjeta)
            # Adiantamento = (valor_total - gorjeta) * 0.6
            valor_total_col = df_final.get('valor_total', 0).fillna(0)
            gorjeta_col = df_final.get('gorjeta', 0).fillna(0)
            df_final['adiantamento_total'] = ((valor_total_col - gorjeta_col) * 0.6).round(2)
            
            # Usar recebedor do consolidado se existir, senão usar do banco
            if 'recebedor' in df_final.columns:
                recebedor_col = df_final['recebedor'].fillna(df_final.get('recebedor_db', ''))
            elif 'recebedor_db' in df_final.columns:
                recebedor_col = df_final['recebedor_db'].fillna('')
            else:
                # Se não tiver recebedor, usar o ID como fallback
                recebedor_col = df_final['id_da_pessoa_entregadora'].fillna('')
            
            # Preparar colunas conforme especificação
            df_excel = pd.DataFrame()
            df_excel['Emissor'] = df_final.get('emissor_db', '').fillna('')
            df_excel['RECEBEDOR'] = recebedor_col.fillna('')
            df_excel['E-MAIL'] = df_final.get('email_db', '').fillna('')
            df_excel['ID'] = df_final['id_da_pessoa_entregadora'].fillna('')
            
            # Valores monetários
            df_excel['Total'] = df_final.get('valor_total', 0).fillna(0)
            df_excel['Nota Fiscal'] = (df_excel['Total'] - df_final.get('gorjeta', 0).fillna(0)).round(2)
            df_excel['Gorjeta'] = df_final.get('gorjeta', 0).fillna(0)
            df_excel['Valor Promocao'] = df_final.get('promo', 0).fillna(0)
            df_excel['Valor Hora Online'] = df_final.get('online_time', 0).fillna(0)
            df_excel['Valor 60%'] = df_final.get('valor_60_percent', 0).fillna(0)
            df_excel['Valor 60% - 0.35'] = df_final.get('valor_final', 0).fillna(0)
            
            # Adiantamentos = 60% do valor total (excluindo gorjeta) do CSV do dia
            # Mesma lógica do "Valor 60%"
            df_excel['Adiantamentos'] = df_excel['Valor 60%']
            # Usar subpracas do consolidado se existir, senão usar subpraca do banco
            if 'subpracas' in df_final.columns:
                subpraca_col = df_final['subpracas'].fillna(df_final.get('subpraca_db', ''))
            else:
                subpraca_col = df_final.get('subpraca_db', '').fillna('')
            df_excel['Subpraça'] = subpraca_col.fillna('')
            df_excel['Notas Emitidas'] = ''
            df_excel['Tipo de Chave Pix'] = df_final.get('tipo_de_chave_pix_db', '').fillna('')
            df_excel['Chave Pix'] = df_final.get('chave_pix_db', '').fillna('')
            df_excel['CNPJ'] = df_final.get('cnpj_db', '').fillna('')
            df_excel['Status'] = ''
            
            # Salvar Excel
            caminho_excel = os.path.join(pasta_relatorios, nome_arquivo)
            df_excel.to_excel(caminho_excel, index=False, engine='openpyxl')
            
            # Formatar Excel
            wb = load_workbook(caminho_excel)
            ws = wb.active
            azul_abjp = "0B5CFF"
            
            # Formatar cabeçalho
            for cell in ws[1]:
                cell.fill = PatternFill(start_color=azul_abjp, end_color=azul_abjp, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Formatar células de dados
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(italic=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(caminho_excel)
            
            flash(f'Excel exportado com sucesso! ({tipo})', 'adiantamento_success')
            return send_file(caminho_excel, as_attachment=True, download_name=nome_arquivo)
            
        except Exception as e:
            flash(f'Erro ao exportar Excel: {str(e)}', 'adiantamento_error')
            return redirect(url_for('lista_solicitacoes'))
